from distutils.cmd import Command
from tkinter import *
from turtle import left
from openpyxl import *

cargararchivo = load_workbook('C:\\Users\\Jose G\\Documents\\scan\\excel.xlsx')  
hoja = cargararchivo.active 
mrow=hoja.max_row
mcol=hoja.max_column

def excel(): 
    hoja.column_dimensions['A'].width = 30
    hoja.column_dimensions['B'].width = 30
    hoja.column_dimensions['C'].width = 30
    hoja.column_dimensions['D'].width = 30
    hoja.column_dimensions['E'].width = 30

    hoja.cell(row=1, column=1).value = "Nombre"
    hoja.cell(row=1, column=2).value = "Número de cedula"
    hoja.cell(row=1, column=3).value = "Número de teléfono"
    hoja.cell(row=1, column=4).value = "Correo electrónico"
    hoja.cell(row=1, column=5).value = "Código de barras"

def importardatos():
    for row in hoja.iter_rows(min_row=1,max_col=5):
        for cell in row:
            if cell==entrada_codigo:
                datos.insert(row[0])


#Interfaz
main=Tk()
main.geometry("1200x650")
main.title("Ingreso de empleados")
main.config(bg="black")
main.resizable(width=FALSE,height=FALSE)
excel()

#Marcos
marco_titulo=Frame(main,width=1200,height=50,bg="black")
marco_centro=Frame(main,width=1200,height=530,bg="light grey")
marco_abajo=Frame(main,width=1200,height=50,bg="black")

marco_titulo.pack(side=TOP)
marco_centro.pack(side=LEFT)
marco_abajo.pack(side=BOTTOM)

#Labels, entrys y cuadros
titulo=Label(marco_titulo,text="Escaner de empleados")
codigo=Label(marco_centro,text="Código de barras")
entrada_codigo=Entry(marco_centro,width=30)
datos=Text(marco_centro,width=75,height=20)

titulo.pack(side=LEFT)
codigo.place(x=3,y=100)
entrada_codigo.place(x=110,y=100)
datos.place(x=400,y=100)

main.mainloop()