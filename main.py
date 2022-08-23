from ast import Str
from ctypes import LibraryLoader
from email import message
from tkinter import *
from tkinter import messagebox
from tokenize import String
from turtle import end_fill, left, width
from openpyxl import *
import time
import random
import tkinter as tk
from PIL import Image,ImageTk


   

main=Tk()
main.resizable(width=FALSE,height=FALSE)
main.geometry("1200x650")
main.title("Punto de venta")


main_marco_arriba=Frame(main,bg="Black",width=1200,height=50)
main_marco_arriba.pack(side=TOP)

main_marco_abajo=Frame(main,bg="Black",width=1200,height=50)
main_marco_abajo.pack(side=BOTTOM)

main_marco_centro=Frame(main,width=1200,height=520)
main_marco_centro.pack(side=LEFT)

icono_facturacion=PhotoImage(file="facturacion.png")
btn_fact=Button(main,text="Facturación",image=icono_facturacion,compound=TOP,command=lambda:facturacion())
btn_fact.place(x=400,y=150)

icono_administracion=PhotoImage(file="administracion.png")
btn_fact=Button(main,text="Administración",image=icono_administracion,compound=TOP)
btn_fact.place(x=400,y=300)

icono_inicio_sesion=PhotoImage(file="inicio_sesion.png")
btn_fact=Button(main,text="Inicio de sesión",image=icono_inicio_sesion,compound=TOP)
btn_fact.place(x=600,y=150)

icono_inventario=PhotoImage(file="Inventario.png")
btn_fact=Button(main,text="Inventario",image=icono_inventario,compound=TOP)
btn_fact.place(x=600,y=300)

def facturacion():
    #FUNCIONES

    #Manejo de erorres
    def manejo_errores(mensaje_error):
        if mensaje_error == "Error":
            messagebox.showerror("Error", "No hay digitos para manejar!")

    #Función para cerrar la aplicación
    def salir():
        punto_venta.destroy()

    #Función para reinicar toda la pantalla para una factura nueva
    def reiniciar_factura():
        try:
            btn_check_var1.set(0)
            btn_check_var2.set(0)
            btn_check_var3.set(0)
            btn_check_var4.set(0)
            btn_check_var5.set(0)
            btn_check_var6.set(0)
            btn_check_var7.set(0)
            btn_check_var8.set(0)
            btn_check_var9.set(0)
            btn_check_var10.set(0)
            btn_check_var11.set(0)
            btn_check_var12.set(0)
            
            otro_producto13.delete(0,END)

            global lista_menu
            menu_opciones_var1.set(lista_menu[0])
            menu_opciones_var2.set(lista_menu[0])
            menu_opciones_var3.set(lista_menu[0])
            menu_opciones_var4.set(lista_menu[0])
            menu_opciones_var5.set(lista_menu[0])
            menu_opciones_var6.set(lista_menu[0])
            menu_opciones_var7.set(lista_menu[0])
            menu_opciones_var8.set(lista_menu[0])
            menu_opciones_var9.set(lista_menu[0])
            menu_opciones_var10.set(lista_menu[0])
            menu_opciones_var11.set(lista_menu[0])
            menu_opciones_var12.set(lista_menu[0])
            menu_opciones_var13.set(lista_menu[0])

            mp2_btn_check_var1.set(0)
            mp2_btn_check_var2.set(0)
            mp2_btn_check_var3.set(0)
            mp2_btn_check_var4.set(0)
            mp2_btn_check_var5.set(0)
            mp2_btn_check_var6.set(0)
            mp2_btn_check_var7.set(0)
            mp2_btn_check_var8.set(0)
            mp2_btn_check_var9.set(0)
            mp2_btn_check_var10.set(0)
            mp2_btn_check_var11.set(0)
            mp2_btn_check_var12.set(0)
            
            mp2_otro_producto13.delete(0,END)

            global mp2_lista_menu
            mp2_menu_opciones_var1.set(mp2_lista_menu[0])
            mp2_menu_opciones_var2.set(mp2_lista_menu[0])
            mp2_menu_opciones_var3.set(mp2_lista_menu[0])
            mp2_menu_opciones_var4.set(mp2_lista_menu[0])
            mp2_menu_opciones_var5.set(mp2_lista_menu[0])
            mp2_menu_opciones_var6.set(mp2_lista_menu[0])
            mp2_menu_opciones_var7.set(mp2_lista_menu[0])
            mp2_menu_opciones_var8.set(mp2_lista_menu[0])
            mp2_menu_opciones_var9.set(mp2_lista_menu[0])
            mp2_menu_opciones_var10.set(mp2_lista_menu[0])
            mp2_menu_opciones_var11.set(mp2_lista_menu[0])
            mp2_menu_opciones_var12.set(mp2_lista_menu[0])
            mp2_menu_opciones_var13.set(mp2_lista_menu[0])

            costo_productos1.delete(0,END)
            costo_productos2.delete(0,END)
            costo_total.delete(0,END)
            impuestos.delete(0,END)
            subtotal.delete(0,END)
            total.delete(0,END)
            espacio_factura.delete("1.0",END)
            precio_cantidad_prods.clear()
            precio_otro_producto13.delete(0,END)
            mp2_precio_otro_producto13.delete(0,END)
            messagebox.showinfo("Siguiente cliente","Atiende al siguiente cliente por favor")
        except NameError:
            messagebox.showerror("Error","Nada por limpiar")

    #Función para mostrar las ventas totales en el resumen
    def ventas_totales():
        precios_prod1={"prod1":20,"prod2":20,"prod3":20,"prod4":20,"prod5":20,"prod6":20,"prod7":20,"prod8":20,"prod9":20,"prod10":20,"prod11":20,"prod12":20,"Otros productos":precio_otro_producto13.get()}

        global precio_cantidad_prods
        precio_cantidad_prods=[]

        if btn_check_var1.get()==1:
            resultado_prod1=int(menu_opciones_var1.get())*precios_prod1["prod1"]
            tupla_prod1=("prod1",str(menu_opciones_var1.get()),"20")
            precio_cantidad_prods.append(tupla_prod1)
        else:
            resultado_prod1=0
        
        if btn_check_var2.get()==1:
            resultado_prod2=int(menu_opciones_var2.get())*precios_prod1["prod2"]
            tupla_prod2=("prod2",str(menu_opciones_var2.get()),"20")
            precio_cantidad_prods.append(tupla_prod2)
        else:
            resultado_prod2=0
        
        if btn_check_var3.get()==1:
            resultado_prod3=int(menu_opciones_var3.get())*precios_prod1["prod3"]
            tupla_prod3=("prod3",str(menu_opciones_var3.get()),"20")
            precio_cantidad_prods.append(tupla_prod3)
        else:
            resultado_prod3=0
        
        if btn_check_var4.get()==1:
            resultado_prod4=int(menu_opciones_var4.get())*precios_prod1["prod4"]
            tupla_prod4=("prod4",str(menu_opciones_var4.get()),"20")
            precio_cantidad_prods.append(tupla_prod4)
        else:
            resultado_prod4=0
        
        if btn_check_var5.get()==1:
            resultado_prod5=int(menu_opciones_var5.get())*precios_prod1["prod5"]
            tupla_prod5=("prod5",str(menu_opciones_var5.get()),"20")
            precio_cantidad_prods.append(tupla_prod5)
        else:
            resultado_prod5=0
        
        if btn_check_var6.get()==1:
            resultado_prod6=int(menu_opciones_var6.get())*precios_prod1["prod6"]
            tupla_prod6=("prod6",str(menu_opciones_var6.get()),"20")
            precio_cantidad_prods.append(tupla_prod6)
        else:
            resultado_prod6=0
        
        if btn_check_var7.get()==1:
            resultado_prod7=int(menu_opciones_var7.get())*precios_prod1["prod7"]
            tupla_prod7=("prod7",str(menu_opciones_var7.get()),"20")
            precio_cantidad_prods.append(tupla_prod7)
        else:
            resultado_prod7=0
        
        if btn_check_var8.get()==1:
            resultado_prod8=int(menu_opciones_var8.get())*precios_prod1["prod8"]
            tupla_prod8=("prod8",str(menu_opciones_var8.get()),"20")
            precio_cantidad_prods.append(tupla_prod8)
        else:
            resultado_prod8=0
        
        if btn_check_var9.get()==1:
            resultado_prod9=int(menu_opciones_var9.get())*precios_prod1["prod9"]
            tupla_prod9=("prod9",str(menu_opciones_var9.get()),"20")
            precio_cantidad_prods.append(tupla_prod9)
        else:
            resultado_prod9=0
        
        if btn_check_var10.get()==1:
            resultado_prod10=int(menu_opciones_var10.get())*precios_prod1["prod10"]
            tupla_prod10=("prod10",str(menu_opciones_var10.get()),"20")
            precio_cantidad_prods.append(tupla_prod10)
        else:
            resultado_prod10=0
        
        if btn_check_var11.get()==1:
            resultado_prod11=int(menu_opciones_var11.get())*precios_prod1["prod11"]
            tupla_prod11=("prod11",str(menu_opciones_var11.get()),"20")
            precio_cantidad_prods.append(tupla_prod11)
        else:
            resultado_prod11=0
        
        if btn_check_var12.get()==1:
            resultado_prod12=int(menu_opciones_var12.get())*precios_prod1["prod12"]
            tupla_prod12=("prod12",str(menu_opciones_var12.get()),"20")
            precio_cantidad_prods.append(tupla_prod12)
        else:
            resultado_prod12=0
        
        try:
            global resultado_otro_prod
            global costo_total_prod1
            if otro_producto13.get():
                resultado_otro_prod=int(menu_opciones_var13.get())*int(precio_otro_producto13.get())
                tupla_otro_prod=(otro_producto13.get(),str(menu_opciones_var13.get()),precio_otro_producto13.get())
                precio_cantidad_prods.append(tupla_otro_prod)
            else:
                resultado_otro_prod=0

            costo_total_prod1= (resultado_prod1+resultado_prod2+resultado_prod3+resultado_prod4+resultado_prod5+resultado_prod6+resultado_prod7+resultado_prod8+resultado_prod9+resultado_prod10+resultado_prod11+resultado_prod12+resultado_otro_prod)

            costo_productos1.delete(0,END)
            costo_productos1.insert(0,"₡ "+str(costo_total_prod1))
        except NameError:
            manejo_errores("Error")
        except ValueError:
            messagebox.showerror("Error","Ponga precio a sus productos primero.")


        precios_prod2={"prod1": 20, "prod2":20,"prod3":20,"prod4":20,"prod5":20,"prod6":20,"prod7":20,"prod8":20,"prod9":20,"prod10":20,"prod11":20,"prod12":20,"Otros productos":mp2_menu_opciones_var13.get()}

        if mp2_btn_check_var1.get()==1:
            mp2_resultado_prod1=int(mp2_menu_opciones_var1.get())*precios_prod2["prod1"]
            mp2_tupla_prod1=("prod1",str(mp2_menu_opciones_var1.get()),"20")
            precio_cantidad_prods.append(mp2_tupla_prod1)
        else:
            mp2_resultado_prod1=0

        if mp2_btn_check_var2.get()==1:
            mp2_resultado_prod2=int(mp2_menu_opciones_var2.get())*precios_prod2["prod2"]
            mp2_tupla_prod2=("prod2",str(mp2_menu_opciones_var2.get()),"20")
            precio_cantidad_prods.append(mp2_tupla_prod2)
        else:
            mp2_resultado_prod2=0
        
        if mp2_btn_check_var3.get()==1:
            mp2_resultado_prod3=int(mp2_menu_opciones_var3.get())*precios_prod2["prod3"]
            mp2_tupla_prod3=("prod3",str(mp2_menu_opciones_var3.get()),"20")
            precio_cantidad_prods.append(mp2_tupla_prod3)
        else:
            mp2_resultado_prod3=0
        
        if mp2_btn_check_var4.get()==1:
            mp2_resultado_prod4=int(mp2_menu_opciones_var4.get())*precios_prod2["prod4"]
            mp2_tupla_prod4=("prod4",str(mp2_menu_opciones_var4.get()),"20")
            precio_cantidad_prods.append(mp2_tupla_prod4)
        else:
            mp2_resultado_prod4=0
        
        if mp2_btn_check_var5.get()==1:
            mp2_resultado_prod5=int(mp2_menu_opciones_var5.get())*precios_prod2["prod5"]
            mp2_tupla_prod5=("prod5",str(mp2_menu_opciones_var5.get()),"20")
            precio_cantidad_prods.append(mp2_tupla_prod5)
        else:
            mp2_resultado_prod5=0
        
        if mp2_btn_check_var6.get()==1:
            mp2_resultado_prod6=int(mp2_menu_opciones_var6.get())*precios_prod2["prod6"]
            mp2_tupla_prod6=("prod6",str(mp2_menu_opciones_var6.get()),"20")
            precio_cantidad_prods.append(mp2_tupla_prod6)
        else:
            mp2_resultado_prod6=0
        
        if mp2_btn_check_var7.get()==1:
            mp2_resultado_prod7=int(mp2_menu_opciones_var7.get())*precios_prod2["prod7"]
            mp2_tupla_prod7=("prod7",str(mp2_menu_opciones_var7.get()),"20")
            precio_cantidad_prods.append(mp2_tupla_prod7)
        else:
            mp2_resultado_prod7=0
        
        if mp2_btn_check_var8.get()==1:
            mp2_resultado_prod8=int(mp2_menu_opciones_var8.get())*precios_prod2["prod8"]
            mp2_tupla_prod8=("prod8",str(mp2_menu_opciones_var8.get()),"20")
            precio_cantidad_prods.append(mp2_tupla_prod8)
        else:
            mp2_resultado_prod8=0
        
        if mp2_btn_check_var9.get()==1:
            mp2_resultado_prod9=int(mp2_menu_opciones_var9.get())*precios_prod2["prod9"]
            mp2_tupla_prod9=("prod9",str(mp2_menu_opciones_var9.get()),"20")
            precio_cantidad_prods.append(mp2_tupla_prod9)
        else:
            mp2_resultado_prod9=0
        
        if mp2_btn_check_var10.get()==1:
            mp2_resultado_prod10=int(mp2_menu_opciones_var10.get())*precios_prod2["prod10"]
            mp2_tupla_prod10=("prod10",str(mp2_menu_opciones_var10.get()),"20")
            precio_cantidad_prods.append(mp2_tupla_prod10)
        else:
            mp2_resultado_prod10=0
        
        if mp2_btn_check_var11.get()==1:
            mp2_resultado_prod11=int(mp2_menu_opciones_var11.get())*precios_prod2["prod11"]
            mp2_tupla_prod11=("prod11",str(mp2_menu_opciones_var11.get()),"20")
            precio_cantidad_prods.append(mp2_tupla_prod11)
        else:
            mp2_resultado_prod11=0
        
        if mp2_btn_check_var12.get()==1:
            mp2_resultado_prod12=int(mp2_menu_opciones_var12.get())*precios_prod2["prod12"]
            mp2_tupla_prod12=("prod12",str(mp2_menu_opciones_var12.get()),"20")
            precio_cantidad_prods.append(mp2_tupla_prod12)
        else:
            mp2_resultado_prod12=0

        try:
            global mp2_resultado_otro_prod
            if mp2_otro_producto13.get():
                mp2_resultado_otro_prod=int(mp2_menu_opciones_var13.get())*int(mp2_precio_otro_producto13.get())
                mp2_tupla_otro_prod=(mp2_otro_producto13.get(),str(mp2_menu_opciones_var13.get()),mp2_precio_otro_producto13.get())
                precio_cantidad_prods.append(mp2_tupla_otro_prod)
            else:
                mp2_resultado_otro_prod=0
            
            global costo_total_prod2
            global impuestos_totales
            costo_total_prod2=(mp2_resultado_prod1+mp2_resultado_prod2+mp2_resultado_prod3+mp2_resultado_prod4+mp2_resultado_prod5+mp2_resultado_prod6+mp2_resultado_prod7+mp2_resultado_prod8+mp2_resultado_prod9+mp2_resultado_prod10+mp2_resultado_prod11+mp2_resultado_prod12+mp2_resultado_otro_prod)

            costo_productos2.delete(0,END)
            costo_productos2.insert(0,"₡ "+str(costo_total_prod2))

            costo_total.delete(0,END)
            costo_total.insert(0,"₡ "+str(costo_total_prod1+costo_total_prod2))

            impuestos_totales=0.13*(costo_total_prod1+costo_total_prod2)

            impuestos.delete(0,END)
            impuestos.insert(0,"₡ "+str(impuestos_totales))

            subtotal.delete(0,END)
            subtotal.insert(0,"₡ "+str(costo_total_prod1+costo_total_prod2))

            total.delete(0,END)
            total.insert(0,"₡ "+str(costo_total_prod1+costo_total_prod2+impuestos_totales))
        except NameError:
            manejo_errores("Error")
        except ValueError:
            messagebox.showerror("Error","Precio primero")

    #Función para mostrar la factura en el espacio asignado
    def imprimir_factura():
        try:
            espacio_factura.delete("1.0",END)
            espacio_factura.insert(END,"          LA TIENDITA"+"\n")
            espacio_factura.insert(END," "+"\n")
            
            x=0
            while True:
                espacio_factura.insert(END,"  "+precio_cantidad_prods[x][0]+" :   "+precio_cantidad_prods[x][1]+" und ₡"+precio_cantidad_prods[x][2]+" c/u"+"\n")
                x+=1

                if x<len(precio_cantidad_prods):
                    continue
                else:
                    break
            
            espacio_factura.insert(END," "+"\n")
            espacio_factura.insert(END,"  Costo productos1   "+costo_productos1_var.get()+"\n")
            espacio_factura.insert(END,"  Costo productos2   "+costo_productos2_var.get()+"\n")
            espacio_factura.insert(END,"  Costo total        "+costo_total_var.get()+"\n")
            espacio_factura.insert(END," "+"\n")
            espacio_factura.insert(END,"***** GRACIAS POR SU VISITA***** "+"\n")
        except IndexError:
            messagebox.showerror("Error","Escoja los productos primero, gracias")
        except NameError:
            messagebox.showerror("Error","Seleccione los productos primero, gracias")

    #Función para guardat los datos de la venta en el libro de excel
    def guardar_datos():
        try:
            global libro
            archivo_destino="Punto_de_venta.xlsx"
            libro=load_workbook(filename=archivo_destino)

            hoja=libro.active

            fila=2
            columna=1
            global hoy 
            hoy=time.strftime("%x")
            id_=random.randint(1,1000000)

            if precio_cantidad_prods:
                while True:
                    if hoja.cell(row=fila,column=columna).value:
                        fila+=1
                        continue
                    else:
                        hoja.cell(row=fila,column=columna).value=hoy
                        columna+=1
                        hoja.cell(row=fila,column=columna).value=id_
                        columna+=1

                        x=len(precio_cantidad_prods)
                        hoja.cell(row=fila,column=columna).value=x
                        columna+=1

                        hoja.cell(row=fila,column=columna).value=costo_total_prod1+costo_total_prod2
                        columna+=1
                        hoja.cell(row=fila,column=columna).value=impuestos_totales
                    
                        if columna==5:
                            break
                libro.save(filename=archivo_destino)
                messagebox.showinfo("Datos guardados","Se han guardado los datos")
            else:
                messagebox.showerror("Error","Nada que agregar")
        except NameError:
            messagebox.showerror("Error","Nada que agregar") 
        
    #Función para generar el reporte de las ventas diarias en base al archivo de excel
    def reporte():
        archivo_destino2="Punto_de_venta.xlsx"
        libro_reporte=load_workbook(filename=archivo_destino2)

        hoja2=libro_reporte.active

        fila=2
        columna=1
        fila_maxima=hoja2.max_row
        monto_total=[]
        impuesto_total=[]
        hoy2=time.strftime("%x")

        x=0
        while x<fila_maxima:
            if hoja2.cell(row=fila,column=columna).value==hoy2:
                monto_total.append(hoja2.cell(row=fila,column=4).value)
                impuesto_total.append(hoja2.cell(row=fila,column=5).value)
                fila+=1
                x+=1
                continue
            else:
                fila+=1
                x+=1
                continue

        respuesta1=0
        for entero in monto_total:
            respuesta1+=entero

        respuesta2=0
        for flotante in impuesto_total:
            respuesta2+=flotante

        espacio_factura.delete("1.0",END)
        espacio_factura.insert(END,"     LA TIENDITA "+hoy2+"."+"\n")
        espacio_factura.insert(END, "  " + "\n")
        espacio_factura.insert(END,      "***REPORTE DIARIO***"+"\n")
        espacio_factura.insert(END, "  " + "\n")
        espacio_factura.insert(END,"Monto total generado hoy:        "+"₡"+str(respuesta1)+"\n")
        espacio_factura.insert(END, "  " + "\n")
        espacio_factura.insert(END,"Impuestos totales generados hoy: "+"₡"+str(respuesta2)+"\n")
        espacio_factura.insert(END, "  " + "\n")
        espacio_factura.insert(END,"Ingresos mas impuestos de hoy:   "+"₡"+str(respuesta1+respuesta2)+"\n")
        espacio_factura.insert(END, "  " + "\n")


    #INTERFAZ

    punto_venta=Toplevel(main)

    #Configuración
    punto_venta.resizable(width=FALSE,height=FALSE)
    punto_venta.geometry("1200x650")
    punto_venta.title("Punto de venta")
    color="Light Gray"
    color2="Gray"
    punto_venta.configure(bg=color)

    #Marcos
    marco_arriba=Frame(punto_venta,bg=color,width=1200, height=50)
    marco_arriba.pack(side=TOP)

    marco_abajo=Frame(punto_venta,bg=color,width=1200,height=150)
    marco_abajo.pack(side=BOTTOM)

    marco_productos1=Frame(punto_venta,bg=color,width=300,height=500)
    marco_productos1.pack(side=LEFT)

    marco_productos2=Frame(punto_venta,bg=color,width=300,height=500)
    marco_productos2.pack(side=LEFT, padx=5)

    marco_resumen=Frame(punto_venta,bg=color,width=800,height=150)
    marco_resumen.place(x=610,y=60)

    marco_factura=Frame(punto_venta,bg=color,width=800,height=200)
    marco_factura.place(x=610,y=210)

    #Titulo
    etiqueta_titulo=Label(marco_arriba,text="La tiendita", font=("Aisha Latin Semibold",20, "bold"),bg=color,pady=10)
    etiqueta_titulo.pack(side=LEFT)

    #Etiquetas y datos de resumen
    etiqueta_costo_productos1=Label(marco_resumen,text="Costo de productos",font=("Aisha Latin Semibold",10,"bold"),bg=color)
    etiqueta_costo_productos1.place(x=10,y=5)

    costo_productos1_var=StringVar()
    costo_productos1=Entry(marco_resumen,textvariable=costo_productos1_var,width=10)
    costo_productos1.place(x=160,y=5)

    etiqueta_costo_productos2=Label(marco_resumen,text="Costo de productos",font=("Aisha Latin Semibold",10,"bold"),bg=color)
    etiqueta_costo_productos2.place(x=10,y=50)

    costo_productos2_var=StringVar()
    costo_productos2=Entry(marco_resumen,textvariable=costo_productos2_var,width=10)
    costo_productos2.place(x=160,y=50)

    etiqueta_costo_total=Label(marco_resumen,text="Costo total",font=("Aisha Latin Semibold",10,"bold"),bg=color)
    etiqueta_costo_total.place(x=10,y=95)

    costo_total_var=StringVar()
    costo_total=Entry(marco_resumen,textvariable=costo_total_var,width=10)
    costo_total.place(x=160,y=95)

    etiqueta_impuestos=Label(marco_resumen,text="Impuestos",font=("Aisha Latin Semibold",10,"bold"),bg=color)
    etiqueta_impuestos.place(x=350,y=5)

    impuestos_var=StringVar()
    impuestos=Entry(marco_resumen,textvariable=impuestos_var,width=10)
    impuestos.place(x=460,y=5)

    etiqueta_subtotal=Label(marco_resumen,text="Subtotal",font=("Aisha Latin Semibold",10,"bold"),bg=color)
    etiqueta_subtotal.place(x=350,y=50)

    subtotal_var=StringVar()
    subtotal=Entry(marco_resumen,textvariable=subtotal_var,width=10)
    subtotal.place(x=460,y=50)

    etiqueta_total=Label(marco_resumen,text="Total",font=("Aisha Latin Semibold",10,"bold"),bg=color)
    etiqueta_total.place(x=350,y=95)

    total_var=StringVar()
    total=Entry(marco_resumen,textvariable=total_var,width=10)
    total.place(x=460,y=95)



    #Espacio de texto para factura
    espacio_factura=Text(marco_factura,width=58,height=14)
    espacio_factura.pack(side=TOP)

    #Checks, etiquetas y cantidades productos 1
    btn_check_var1=IntVar()
    btn_check1=Checkbutton(marco_productos1,text="prod1",variable=btn_check_var1)
    btn_check1.place(x=3,y=5)

    btn_check_var2=IntVar()
    btn_check2=Checkbutton(marco_productos1,text="prod2",variable=btn_check_var2)
    btn_check2.place(x=3,y=35)

    btn_check_var3=IntVar()
    btn_check3=Checkbutton(marco_productos1,text="prod3",variable=btn_check_var3)
    btn_check3.place(x=3,y=65)

    btn_check_var4=IntVar()
    btn_check4=Checkbutton(marco_productos1,text="prod4",variable=btn_check_var4)
    btn_check4.place(x=3,y=95)

    btn_check_var5=IntVar()
    btn_check5=Checkbutton(marco_productos1,text="prod5",variable=btn_check_var5)
    btn_check5.place(x=3,y=125)

    btn_check_var6=IntVar()
    btn_check6=Checkbutton(marco_productos1,text="prod6",variable=btn_check_var6)
    btn_check6.place(x=3,y=155)

    btn_check_var7=IntVar()
    btn_check7=Checkbutton(marco_productos1,text="prod7",variable=btn_check_var7)
    btn_check7.place(x=3,y=185)

    btn_check_var8=IntVar()
    btn_check8=Checkbutton(marco_productos1,text="prod8",variable=btn_check_var8)
    btn_check8.place(x=3,y=215)

    btn_check_var9=IntVar()
    btn_check9=Checkbutton(marco_productos1,text="prod9",variable=btn_check_var9)
    btn_check9.place(x=3,y=245)

    btn_check_var10=IntVar()
    btn_check10=Checkbutton(marco_productos1,text="prod10",variable=btn_check_var10)
    btn_check10.place(x=3,y=275)

    btn_check_var11=IntVar()
    btn_check11=Checkbutton(marco_productos1,text="prod11",variable=btn_check_var11)
    btn_check11.place(x=3,y=305)

    btn_check_var12=IntVar()
    btn_check12=Checkbutton(marco_productos1,text="prod12",variable=btn_check_var12)
    btn_check12.place(x=3,y=335)

    etiqueta_otro_producto13=Label(marco_productos1,text="Otros productos")
    etiqueta_otro_producto13.place(x=3,y=390)

    otro_producto13=Entry(marco_productos1,width=10)
    otro_producto13.place(x=132,y=390)

    etiqueta_precio_otro_producto13=Label(marco_productos1,text="Precio otros productos")
    etiqueta_precio_otro_producto13.place(x=3,y=420)

    precio_otro_producto13=Entry(marco_productos1,width=10)
    precio_otro_producto13.place(x=132,y=420)


    #Opciones del menu productos 1
    lista_menu=["0","1","2","3","4","5","6","7","8","9","10"]


    menu_opciones_var1=StringVar()
    menu_opciones_var1.set(lista_menu[0])
    menu_opciones1=OptionMenu(marco_productos1,menu_opciones_var1,*lista_menu)
    menu_opciones1.configure(bg=color,font=("arial",7))
    menu_opciones1.place(x=200,y=5)

    menu_opciones_var2=StringVar()
    menu_opciones_var2.set(lista_menu[0])
    menu_opciones2=OptionMenu(marco_productos1,menu_opciones_var2,*lista_menu)
    menu_opciones2.configure(bg=color,font=("arial",7))
    menu_opciones2.place(x=200,y=35)

    menu_opciones_var3=StringVar()
    menu_opciones_var3.set(lista_menu[0])
    menu_opciones3=OptionMenu(marco_productos1,menu_opciones_var3,*lista_menu)
    menu_opciones3.configure(bg=color,font=("arial",7))
    menu_opciones3.place(x=200,y=65)

    menu_opciones_var4=StringVar()
    menu_opciones_var4.set(lista_menu[0])
    menu_opciones4=OptionMenu(marco_productos1,menu_opciones_var4,*lista_menu)
    menu_opciones4.configure(bg=color,font=("arial",7))
    menu_opciones4.place(x=200,y=95)

    menu_opciones_var5=StringVar()
    menu_opciones_var5.set(lista_menu[0])
    menu_opciones5=OptionMenu(marco_productos1,menu_opciones_var5,*lista_menu)
    menu_opciones5.configure(bg=color,font=("arial",7))
    menu_opciones5.place(x=200,y=125)

    menu_opciones_var6=StringVar()
    menu_opciones_var6.set(lista_menu[0])
    menu_opciones6=OptionMenu(marco_productos1,menu_opciones_var6,*lista_menu)
    menu_opciones6.configure(bg=color,font=("arial",7))
    menu_opciones6.place(x=200,y=155)

    menu_opciones_var7=StringVar()
    menu_opciones_var7.set(lista_menu[0])
    menu_opciones7=OptionMenu(marco_productos1,menu_opciones_var7,*lista_menu)
    menu_opciones7.configure(bg=color,font=("arial",7))
    menu_opciones7.place(x=200,y=185)

    menu_opciones_var8=StringVar()
    menu_opciones_var8.set(lista_menu[0])
    menu_opciones8=OptionMenu(marco_productos1,menu_opciones_var8,*lista_menu)
    menu_opciones8.configure(bg=color,font=("arial",7))
    menu_opciones8.place(x=200,y=215)

    menu_opciones_var9=StringVar()
    menu_opciones_var9.set(lista_menu[0])
    menu_opciones9=OptionMenu(marco_productos1,menu_opciones_var9,*lista_menu)
    menu_opciones9.configure(bg=color,font=("arial",7))
    menu_opciones9.place(x=200,y=245)

    menu_opciones_var10=StringVar()
    menu_opciones_var10.set(lista_menu[0])
    menu_opciones10=OptionMenu(marco_productos1,menu_opciones_var10,*lista_menu)
    menu_opciones10.configure(bg=color,font=("arial",7))
    menu_opciones10.place(x=200,y=275)

    menu_opciones_var11=StringVar()
    menu_opciones_var11.set(lista_menu[0])
    menu_opciones11=OptionMenu(marco_productos1,menu_opciones_var11,*lista_menu)
    menu_opciones11.configure(bg=color,font=("arial",7))
    menu_opciones11.place(x=200,y=305)

    menu_opciones_var12=StringVar()
    menu_opciones_var12.set(lista_menu[0])
    menu_opciones12=OptionMenu(marco_productos1,menu_opciones_var12,*lista_menu)
    menu_opciones12.configure(bg=color,font=("arial",7))
    menu_opciones12.place(x=200,y=335)

    menu_opciones_var13=StringVar()
    menu_opciones_var13.set(lista_menu[0])
    menu_opciones13=OptionMenu(marco_productos1,menu_opciones_var13,*lista_menu)
    menu_opciones13.configure(bg=color,font=("arial",7))
    menu_opciones13.place(x=200,y=390)


    #Checks, etiquetas y cantidades productos 2
    mp2_btn_check_var1=IntVar()
    mp2_btn_check1=Checkbutton(marco_productos2,text="prod1",variable=mp2_btn_check_var1)
    mp2_btn_check1.place(x=3,y=5)

    mp2_btn_check_var2=IntVar()
    mp2_btn_check2=Checkbutton(marco_productos2,text="prod2",variable=mp2_btn_check_var2)
    mp2_btn_check2.place(x=3,y=35)

    mp2_btn_check_var3=IntVar()
    mp2_btn_check3=Checkbutton(marco_productos2,text="prod3",variable=mp2_btn_check_var3)
    mp2_btn_check3.place(x=3,y=65)

    mp2_btn_check_var4=IntVar()
    mp2_btn_check4=Checkbutton(marco_productos2,text="prod4",variable=mp2_btn_check_var4)
    mp2_btn_check4.place(x=3,y=95)

    mp2_btn_check_var5=IntVar()
    mp2_btn_check5=Checkbutton(marco_productos2,text="prod5",variable=mp2_btn_check_var5)
    mp2_btn_check5.place(x=3,y=125)

    mp2_btn_check_var6=IntVar()
    mp2_btn_check6=Checkbutton(marco_productos2,text="prod6",variable=mp2_btn_check_var6)
    mp2_btn_check6.place(x=3,y=155)

    mp2_btn_check_var7=IntVar()
    mp2_btn_check7=Checkbutton(marco_productos2,text="prod7",variable=mp2_btn_check_var7)
    mp2_btn_check7.place(x=3,y=185)

    mp2_btn_check_var8=IntVar()
    mp2_btn_check8=Checkbutton(marco_productos2,text="prod8",variable=mp2_btn_check_var8)
    mp2_btn_check8.place(x=3,y=215)

    mp2_btn_check_var9=IntVar()
    mp2_btn_check9=Checkbutton(marco_productos2,text="prod9",variable=mp2_btn_check_var9)
    mp2_btn_check9.place(x=3,y=245)

    mp2_btn_check_var10=IntVar()
    mp2_btn_check10=Checkbutton(marco_productos2,text="prod10",variable=mp2_btn_check_var10)
    mp2_btn_check10.place(x=3,y=275)

    mp2_btn_check_var11=IntVar()
    mp2_btn_check11=Checkbutton(marco_productos2,text="prod11",variable=mp2_btn_check_var11)
    mp2_btn_check11.place(x=3,y=305)

    mp2_btn_check_var12=IntVar()
    mp2_btn_check12=Checkbutton(marco_productos2,text="prod12",variable=mp2_btn_check_var12)
    mp2_btn_check12.place(x=3,y=335)

    mp2_etiqueta_otro_producto13=Label(marco_productos2,text="Otros productos")
    mp2_etiqueta_otro_producto13.place(x=3,y=390)

    mp2_otro_producto13=Entry(marco_productos2,width=10)
    mp2_otro_producto13.place(x=132,y=390)

    mp2_etiqueta_precio_otro_producto13=Label(marco_productos2,text="Precio otros productos")
    mp2_etiqueta_precio_otro_producto13.place(x=3,y=420)

    mp2_precio_otro_producto13=Entry(marco_productos2,width=10)
    mp2_precio_otro_producto13.place(x=132,y=420)

    #Opciones del menu productos 2
    mp2_lista_menu=["0","1","2","3","4","5","6","7","8","9","10"]


    mp2_menu_opciones_var1=StringVar()
    mp2_menu_opciones_var1.set(mp2_lista_menu[0])
    mp2_menu_opciones1=OptionMenu(marco_productos2,mp2_menu_opciones_var1,*mp2_lista_menu)
    mp2_menu_opciones1.configure(bg=color,font=("arial",7))
    mp2_menu_opciones1.place(x=200,y=5)

    mp2_menu_opciones_var2=StringVar()
    mp2_menu_opciones_var2.set(mp2_lista_menu[0])
    mp2_menu_opciones2=OptionMenu(marco_productos2,mp2_menu_opciones_var2,*mp2_lista_menu)
    mp2_menu_opciones2.configure(bg=color,font=("arial",7))
    mp2_menu_opciones2.place(x=200,y=35)

    mp2_menu_opciones_var3=StringVar()
    mp2_menu_opciones_var3.set(mp2_lista_menu[0])
    mp2_menu_opciones3=OptionMenu(marco_productos2,mp2_menu_opciones_var3,*mp2_lista_menu)
    mp2_menu_opciones3.configure(bg=color,font=("arial",7))
    mp2_menu_opciones3.place(x=200,y=65)

    mp2_menu_opciones_var4=StringVar()
    mp2_menu_opciones_var4.set(mp2_lista_menu[0])
    mp2_menu_opciones4=OptionMenu(marco_productos2,mp2_menu_opciones_var4,*mp2_lista_menu)
    mp2_menu_opciones4.configure(bg=color,font=("arial",7))
    mp2_menu_opciones4.place(x=200,y=95)

    mp2_menu_opciones_var5=StringVar()
    mp2_menu_opciones_var5.set(mp2_lista_menu[0])
    mp2_menu_opciones5=OptionMenu(marco_productos2,mp2_menu_opciones_var5,*mp2_lista_menu)
    mp2_menu_opciones5.configure(bg=color,font=("arial",7))
    mp2_menu_opciones5.place(x=200,y=125)

    mp2_menu_opciones_var6=StringVar()
    mp2_menu_opciones_var6.set(mp2_lista_menu[0])
    mp2_menu_opciones6=OptionMenu(marco_productos2,mp2_menu_opciones_var6,*mp2_lista_menu)
    mp2_menu_opciones6.configure(bg=color,font=("arial",7))
    mp2_menu_opciones6.place(x=200,y=155)

    mp2_menu_opciones_var7=StringVar()
    mp2_menu_opciones_var7.set(mp2_lista_menu[0])
    mp2_menu_opciones7=OptionMenu(marco_productos2,mp2_menu_opciones_var7,*mp2_lista_menu)
    mp2_menu_opciones7.configure(bg=color,font=("arial",7))
    mp2_menu_opciones7.place(x=200,y=185)

    mp2_menu_opciones_var8=StringVar()
    mp2_menu_opciones_var8.set(mp2_lista_menu[0])
    mp2_menu_opciones8=OptionMenu(marco_productos2,mp2_menu_opciones_var8,*mp2_lista_menu)
    mp2_menu_opciones8.configure(bg=color,font=("arial",7))
    mp2_menu_opciones8.place(x=200,y=215)

    mp2_menu_opciones_var9=StringVar()
    mp2_menu_opciones_var9.set(mp2_lista_menu[0])
    mp2_menu_opciones9=OptionMenu(marco_productos2,mp2_menu_opciones_var9,*mp2_lista_menu)
    mp2_menu_opciones9.configure(bg=color,font=("arial",7))
    mp2_menu_opciones9.place(x=200,y=245)

    mp2_menu_opciones_var10=StringVar()
    mp2_menu_opciones_var10.set(mp2_lista_menu[0])
    mp2_menu_opciones10=OptionMenu(marco_productos2,mp2_menu_opciones_var10,*mp2_lista_menu)
    mp2_menu_opciones10.configure(bg=color,font=("arial",7))
    mp2_menu_opciones10.place(x=200,y=275)

    mp2_menu_opciones_var11=StringVar()
    mp2_menu_opciones_var11.set(mp2_lista_menu[0])
    mp2_menu_opciones11=OptionMenu(marco_productos2,mp2_menu_opciones_var11,*mp2_lista_menu)
    mp2_menu_opciones11.configure(bg=color,font=("arial",7))
    mp2_menu_opciones11.place(x=200,y=305)

    mp2_menu_opciones_var12=StringVar()
    mp2_menu_opciones_var12.set(mp2_lista_menu[0])
    mp2_menu_opciones12=OptionMenu(marco_productos2,mp2_menu_opciones_var12,*mp2_lista_menu)
    mp2_menu_opciones12.configure(bg=color,font=("arial",7))
    mp2_menu_opciones12.place(x=200,y=335)

    mp2_menu_opciones_var13=StringVar()
    mp2_menu_opciones_var13.set(mp2_lista_menu[0])
    mp2_menu_opciones13=OptionMenu(marco_productos2,mp2_menu_opciones_var13,*mp2_lista_menu)
    mp2_menu_opciones13.configure(bg=color,font=("arial",7))
    mp2_menu_opciones13.place(x=200,y=390)


    #Botones funcionales
    btn_total=Button(marco_abajo,text="Total",width=10,bg=color2,command=lambda:ventas_totales())
    btn_total.place(x=3,y=10)

    btn_factura=Button(marco_abajo,text="Factura",width=10,bg=color2,command=lambda:imprimir_factura())
    btn_factura.place(x=100,y=10)

    btn_reiniciar=Button(marco_abajo,text="Reiniciar",width=10, bg=color2,command=lambda:reiniciar_factura())
    btn_reiniciar.place(x=200,y=10)

    btn_guardar_datos=Button(marco_abajo,text="Guardar datos",width=10,bg=color2,command=lambda:guardar_datos())
    btn_guardar_datos.place(x=3,y=50)

    btn_reporte=Button(marco_abajo,text="Reporte",width=10,bg=color2,command=lambda:reporte())
    btn_reporte.place(x=100,y=50)

    btn_salir=Button(marco_abajo,text="Salir",width=10,bg=color2,command=lambda:salir())
    btn_salir.place(x=200,y=50)

    punto_venta.mainloop()



main.mainloop()




    